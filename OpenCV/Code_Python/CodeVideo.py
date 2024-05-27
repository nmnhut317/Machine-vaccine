'''
A simple Program for grabing video from basler camera and converting it to opencv img.
Tested on Basler acA1300-200uc (USB3, linux 64bit , python 3.5)

'''
from pypylon import pylon
import cv2 as cv
import numpy as np
import glob
import pandas as pd
import time
import os
from datetime import datetime
import serial
import threading
import queue
import matplotlib.pyplot as plt
from openpyxl import Workbook

# ser = serial.Serial('COM3', 115200)
wb = Workbook()
ws = wb.active
ws.append(['ID','LENGTH', 'INJECTION', 'TIME', 'IMAGES'])

n = 0
id = 0
collums = 2
scale = 0.5

lengthFish = 0
positionCurrent = 0
positionAhead = 0
# verifyPosition = False
currentTime = datetime.now()

pathSaveImage = 'd'
pathSaveData = "D:/Project/datas.xlsx"


def click_event(event, x, y, flags, param):
    if event == cv.EVENT_RBUTTONDOWN:
        pass


def nothing(x):
        pass


CHECK_DIR = os.path.isdir(pathSaveImage)

# if directory does not exist create
if not CHECK_DIR:
    os.makedirs(pathSaveImage)
    print(f'"{pathSaveImage}" Directory is created')
else:
    print(f'"{pathSaveImage}" Directory already Exists.') 


def image_resize(image, width = None, height = None, inter = cv.INTER_AREA):
    # initialize the dimensions of the image to be resized and
    # grab the image size
    dim = None
    (h, w) = image.shape[:2]

    # if both the width and height are None, then return the
    # original image
    if width is None and height is None:
        return image

    # check to see if the width is None
    if width is None:
        # calculate the ratio of the height and construct the
        # dimensions
        r = height / float(h)
        dim = (int(w * r), height)

    # otherwise, the height is None
    else:
        # calculate the ratio of the width and construct the
        # dimensions
        r = width / float(w)
        dim = (width, int(h * r))

    # resize the image
    resized = cv.resize(image, dim, interpolation = inter)

    # return the resized image
    return resized


def drawCenterLine(img2Draw, binaryImage):
    point = findPoints(binaryImage)
    site = [0, 0.10, 0.20, 0.28, 0.35, 0.5, 0.85, 1]
    pointSite = []
    siteLength = len(point)

    for i in range(siteLength):
        t = int(site[i]*img2Draw.shape[1] - 2)

        if t < 0:
            t = 0

        pointSite.append(point[t][2])

        if site[i] == 1:
            cv.circle(img2Draw, point[t][2], 2, (0,0,255), -1)
            break

        cv.line(img2Draw, point[t][1][0], point[t][1][1], (255,255,50), 2)
        cv.circle(img2Draw, point[t][2], 2, (0,0,255), -1)

    cv.polylines(img2Draw, [np.array(pointSite)], isClosed= False, color= (255, 255, 50), thickness= 2) 

    return pointSite, point


def distanceMultiplePoints(arrayPoint):

    datatype = object
    arrayPoint = np.array(arrayPoint, dtype= datatype)
    # print("after length",arrayPoint, len(arrayPoint))
    # priviousPoint = (arrayPoint[0][0], arrayPoint[0][1])
    priviousPoint = arrayPoint[0]
    dist = 0
    numOfArray = len(arrayPoint)
    # print("length", numOfArray)

    for i in range(numOfArray):
        currentPoint = arrayPoint[i]
        dist = dist + np.linalg.norm(priviousPoint - currentPoint)
        priviousPoint = currentPoint

    dist = round(dist, 1)

    return dist


def convertLenghToInjection(length):
    positionInjection = int((length*0.3362 + 2.6474)*0.63)
    return positionInjection


def isCheckDirection(positionCurrent, positionAhead):
    if positionCurrent < positionAhead :
        positionDynamic = positionAhead - positionCurrent
        transmitData(0, positionDynamic) # 0 tien

    elif positionCurrent > positionAhead:
        positionDynamic = abs(positionAhead - positionCurrent)
        transmitData(1, positionDynamic) # 1 lui

    else:
        pass     


def transmitData(direction, position):
    value = bytes('{} {}\n'.format(direction, position), encoding= 'utf8')
    # ser.write(value)


def calibInputImage(img, cameraMatrix, Distortion_Parameters, path, id):
    h, w = img.shape[:2]
    newcameramtx, roi = cv.getOptimalNewCameraMatrix(cameraMatrix, Distortion_Parameters, (w,h), 1, (w,h))
    # undistort
    Distortion = cv.undistort(img, cameraMatrix, Distortion_Parameters, None, newcameramtx)
    # crop the image
    x, y, w, h = roi
    Distortion = Distortion[y:y+h, x:x+w]
    # cv.imwrite('{}/calibresult{}.jpg'.format(path, id), Distortion)
    # outputImage= cv.imread('{}/calibresult{}.jpg'.format(path, id))

    return Distortion


def findContour(threshold):
    gray = cv.cvtColor(threshold, cv.COLOR_BGR2GRAY)
    threshold = cv.threshold(gray, 0, 255, cv.THRESH_BINARY)[1]
    contour = cv.findContours(threshold, cv.RETR_EXTERNAL, cv.CHAIN_APPROX_SIMPLE)[0]

    if len(contour)>0:
        contour = max(contour, key= cv.contourArea) 
        rect = cv.minAreaRect(contour)

    return rect 


def processImageByCenterLine(readImage, path, id, printValue= False):
    global positionCurrent, positionAhead, collums, scale
    
    img= calibInputImage(readImage, cameraMatrix, Distortion_Parameters, path, id)
    newBinaryImg= np.zeros(img.shape, dtype= np.uint8)

    hsv = cv.cvtColor(img, cv.COLOR_BGR2HSV)
    low_hsv = np.array([65, 110, 104])
    high_hsv = np.array([81, 255, 255])

    mask = cv.inRange(hsv, low_hsv, high_hsv)
    mask = ~mask
    kernel = np.ones((5,5),np.uint8)
    mask = cv.erode(mask, kernel)
    mask = cv.medianBlur(mask, 5)
    # mask = cv.morphologyEx(mask, cv.MORPH_CLOSE, kernel, iterations= 1)
    contour = cv.findContours(mask, cv.RETR_EXTERNAL, cv.CHAIN_APPROX_SIMPLE)[0]

    if len(contour) > 0:
        contour = max(contour, key= cv.contourArea) 

        #	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#
        #                   DRAW A NEW IMAGE            #
        cv.drawContours(newBinaryImg, [contour], 0, (255, 255, 255), -1)
        # cv.imwrite('{}\\binary{}.jpg'.format(path, id), newBinaryImg) # 11/18
        rect = findContour(newBinaryImg)

        #	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#
        #222, 1283
        subBinaryImg = getSubImage(rect, newBinaryImg, img)[0]
        # cv.imwrite('{}\\subbinary{}.jpg'.format(path, id), subBinaryImg) # 11/18
        newMask = cv.split(newBinaryImg)[0]
        result = cv.bitwise_and(img, img, mask= newMask)

        subColor =  getSubImage(rect, result, img)[0]


        subCenterLine = subColor.copy()
        pointSite, onsitePoints = drawCenterLine(subCenterLine, subBinaryImg)
        # print("len onsite = ", len(onsitePoints))

        dist = distanceMultiplePoints(pointSite)
        print(f"\nlength pixel= {dist}(pixel)")
        lengthFish = calculatorLengthByCenterLine((1/scale)*(dist+3), writeLength= True)

        cv.imwrite('{}\\newImage{}.jpg'.format(path, id), result)
        cv.imwrite('{}\\newConvertImage{}.jpg'.format(path, id), subColor)

        drawInjectionImage = subColor.copy()
        # print("shape width: ", drawInjectionImage.shape[1])

        back, belly, midLine, check, splitBellyBinary = isDetermineBackAndBelly(subColor, subBinaryImg, onsitePoints, onsite= True)

        borderImage = None

        if check == True:
            #	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	# 

            positionCurrent = convertLenghToInjection(lengthFish) # VI TRI TIEM MM
            injection_pixel = int((positionCurrent / 0.17)/(1/scale)) # VI TRI TIEM PIXEL

            isCheckDirection(positionCurrent, positionAhead)
            positionAhead = positionCurrent

            print(f'\ninjection at = {positionCurrent}(mm)')
                  
            # print(f'\ninjection at = {positionCurrent}(mm) the same with {injection_pixel}(pixel)')
            
            injection_OnImgX = int(drawInjectionImage.shape[1] - injection_pixel)
            injection_OnImgY = drawInjectionImage.shape[0]

            # print(f"\nshape subColor = {drawInjectionImage.shape}, poisition at ({injection_OnImgX}, {injection_OnImgY})")
            
            # print(f'(x, y) = ({injection_OnImgX}, {onsitePoints[injection_OnImgX][1][1][1]})')

            cv.circle(drawInjectionImage, (injection_OnImgX, onsitePoints[injection_OnImgX][1][1][1]), 10, (0,0,255), 2)
            borderImage = cv.copyMakeBorder(drawInjectionImage, 50, 100, 50, 50, cv.BORDER_CONSTANT, value=[0,0,0]) 
            cv.putText(borderImage, f'Injection position={positionCurrent}mm', (20, borderImage.shape[0] - 65), cv.FONT_HERSHEY_COMPLEX, 1, (255,255,255))
            cv.imwrite('{}\\drawInjectionImage{}.jpg'.format(path, id), drawInjectionImage)
            #	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#

        print("time process= {} ms\n".format(int((time.perf_counter_ns() - start)/1000000)))
        
        # borderImage = cv.copyMakeBorder(drawInjectionImage, 50, 50, 50, 50, cv.BORDER_CONSTANT, value=[0,0,0]) 
        cv.putText(borderImage, f'Length fish={lengthFish}mm', (20, borderImage.shape[0] - 20), cv.FONT_HERSHEY_COMPLEX, 1, (255,255,255))
        cv.imwrite('{}\\borderImage{}.jpg'.format(path, id), borderImage)

        # write data into excel
        ws[f'A{collums}'].value = id
        ws[f'B{collums}'].value = lengthFish
        ws[f'C{collums}'].value = positionCurrent
        ws[f'D{collums}'].value = currentTime
        collums += 1
        wb.save(pathSaveData)

        # colorBelly = splitFin(belly.copy(), splitBellyBinary)

        result = cv.resize(result, (0,0), fx=0.5, fy=0.5)
        subColor = cv.resize(subColor, (0,0), fx= 0.5, fy= 0.5)

        mergeB2B = np.hstack((back, belly))
        mergeB2B = cv.resize(mergeB2B, (0,0), fx= 0.5, fy= 0.5)
        mergeCenter = np.hstack((midLine, subCenterLine, drawInjectionImage))
        mergeCenter = cv.resize(mergeCenter, (0,0), fx= 0.5, fy= 0.5)

        # cv.imwrite('{}\\colorBelly{}.jpg'.format(path, id), colorBelly)
        cv.imwrite('{}\\subCenterLine{}.jpg'.format(path, id), subCenterLine)
        # cv.imwrite('{}\\midLine{}.jpg'.format(path, id), midLine)

        cv.imshow("new image", result)
        # cv.imshow("back and belly", mergeB2B)
        cv.imshow("subColor", mergeCenter)
        cv.imshow("borderImage", borderImage)
        # cv.imshow("subimage", subCenterLine)


def findThreshold(img):
    contour = cv.findContours(img, cv.RETR_EXTERNAL, cv.CHAIN_APPROX_SIMPLE)[0]

    if len(contour) > 0:
        contour = sorted(contour, key= cv.contourArea, reverse= True)
        contour = contour[0]
        rects = cv.minAreaRect(contour) 

    return rects


def saveImage(drawframe, img, threshold, path, takeImage = False):
    global id, start, cameraMatrix, Distortion_Parameters

    contour = cv.findContours(threshold, cv.RETR_EXTERNAL, cv.CHAIN_APPROX_SIMPLE)[0]
    
    contour = max(contour, key= cv.contourArea)

    if len(contour) > 0:
            x, y, w, h = cv.boundingRect(contour)
            CX = int((x+w)/2)
            CY = int((y+h)/2)

            if cv.contourArea(contour) > 20000 and takeImage == True:
                cv.circle(drawframe, (CX , CY), 2, (0,0,255), 15)
                cv.rectangle(drawframe, (x,y), (x+w, y+h), (0,0,255), 1)
                # cv.putText(drawframe, 'CX= {} CY= {}'.format(CX , CY), (CX , CY), cv.FONT_HERSHEY_COMPLEX, 1, 255, 2)
                if CX > int(img.shape[1]/2) - 120  and CX < int(img.shape[1]/2) - 90:
                    start = time.perf_counter_ns()
                    id += 1                           
                    processImageByCenterLine(img, path, id, True)


def calculatorLengthByCenterLine(lengthPixel, lengthFrame = 340, widthPixelFrame = 1920, writeLength = False):

    length = int(lengthPixel* 0.17)

    if writeLength == True:
        print(f"length = {int(length)}(mm)")

    return length


def calibPylon(cameraMatrix, dist, img):
    # Undistortion  
    h, w = img.shape[:2]
    newcameramtx, roi = cv.getOptimalNewCameraMatrix(cameraMatrix, dist, (w,h), 1, (w,h))
    # undistort
    dst = cv.undistort(img, cameraMatrix, dist, None, newcameramtx)

    return dst


def determineLocateLine(img):
    cv.line(img, (int(img.shape[1]/2) - 90, 0), (int(img.shape[1]/2) - 90, img.shape[0]), (0,0,255), 5)
    cv.line(img, (int(img.shape[1]/2) - 120, 0), (int(img.shape[1]/2) - 120, img.shape[0]), (0,255,0), 5)


def getSubImage(rect, src, img, drawPolyline = False):
        
    box = cv.boxPoints(rect)
    box = np.intp(box)
        
    width = int(rect[1][0])
    height = int(rect[1][1])

    
    theta = rect[2]

    if drawPolyline == True:
        cv.polylines(img, [box], True, (0,0,255), 3)

    src_pts = box.astype("float32")
    dst_pts = np.array([[0, height-1],
                                [0, 0],
                                [width-1, 0],
                                [width-1, height-1]], dtype="float32")

    M = cv.getPerspectiveTransform(src_pts, dst_pts)
    warped = cv.warpPerspective(src, M, (width, height))

    if width < height :
        warped = cv.rotate(warped, cv.ROTATE_90_CLOCKWISE) 
           
    return warped, width  


def calibCamera(chessboardSize, frameSize, pathImage, saveImage = False):

    # termination criteria
    criteria = (cv.TERM_CRITERIA_EPS + cv.TERM_CRITERIA_MAX_ITER, 30, 0.001)

    # prepare object points, like (0,0,0), (1,0,0), (2,0,0) ....,(6,5,0)
    objp = np.zeros((chessboardSize[0]*chessboardSize[1], 3), np.float32)
    objp[:,:2] = np.mgrid[0:chessboardSize[0], 0: chessboardSize[1]].T.reshape(-1,2)

    # Arrays to store object points and image points from all the images.
    objpoints = [] # 3d point in real world space
    imgpoints = [] # 2d points in image plane.

    images = glob.glob('{}/*.jpg'.format(str(pathImage)))

    a_number_of_images = 1

    for fname in images:

        img = cv.imread(fname)

        gray = cv.cvtColor(img, cv.COLOR_BGR2GRAY)

        # Find the chess board corners
        ret, corners = cv.findChessboardCorners(gray, chessboardSize, None)

        # If found, add object points, image points (after refining them)
        if ret == True:
            objpoints.append(objp)
            corners2 = cv.cornerSubPix(gray,corners, (11,11), (-1,-1), criteria)
            imgpoints.append(corners2)

            # Draw and display the corners
            cv.drawChessboardCorners(img, chessboardSize, corners2, ret)

            h, w = img.shape[:2]
            print(img.shape)
            img = cv.resize(img, (int(w/2), int(h/2)), fx= 0.5, fy= 0.5)
            
            # cv.imshow('img', img)

            # cv.waitKey(1000)

        a_number_of_images += 1

    print('a_number_of_images:', a_number_of_images)

    cv.destroyAllWindows()

    # calibration
    ret, cameraMatrix, dist, rvecs, tvecs = cv.calibrateCamera(objpoints, imgpoints, frameSize, None, None)
    # print("Camera Calibrate:", ret)
    print("\ncameraMatrix", cameraMatrix)
    print("\nDistortion Parameters:", dist)
    # print("\nRotation Vector:", rvecs)
    # print("\nTranstation Vector:", tvecs)

    # pickle.dump((cameraMatrix, dist), open('calibration.pkl', "wb"))
    # pickle.dump(cameraMatrix, open('cameraMatrix.pkl', "wb"))
    # pickle.dump(dist, open('calibration.pkl', "wb"))

    k = 1
    #               the same feature            #
    # THE FIRST METHOD
    if saveImage == True:
        # Undistortion  
        img = cv.imread('{}\Image1.png'.format(pathImage))
        h, w = img.shape[:2]
        newcameramtx, roi = cv.getOptimalNewCameraMatrix(cameraMatrix, dist, (w,h), 1, (w,h))
        # undistort
        dst = cv.undistort(img, cameraMatrix, dist, None, newcameramtx)
        # crop the image
        x, y, w, h = roi
        dst = dst[y:y+h, x:x+w]
        cv.imwrite('{}\calibresult{}.png'.format(pathImage, k), dst)
        k += 1

    # check error
    mean_error = 0

    for i in range(len(objpoints)):
            imgpoints2, _ = cv.projectPoints(objpoints[i], rvecs[i], tvecs[i], cameraMatrix, dist)
            error = cv.norm(imgpoints[i], imgpoints2, cv.NORM_L2)/len(imgpoints2)
            mean_error += error

    print( "total error: {}".format(mean_error/len(objpoints)) ) 
        
    return cameraMatrix, dist


def findPoints(img, indxStart = 0, indxEnd = 1):
    # start1 = time.perf_counter_ns()
    # find smalles value in each column
    origi = int(img.shape[1])
    midPt = int(img.shape[1] / 2)
    firstPt = int(img.shape[1] /7)
    limitPt = int(img.shape[1] - img.shape[1] /10)
    coordinate = []
    disPoint = []
    points = []

    indxStart= int(origi*indxStart)
    indxEnd = int(origi*indxEnd)


    # for i in range(img.shape[1]):
    for i in range(indxStart, indxEnd):
    # for i in range(midPt, limitPt):
        tmp_col = np.where((img[:, i] == 255))

        if len(tmp_col[0]) > 2:
            firstPoint =  (i, tmp_col[0][0])
            lastPoint =  (i, tmp_col[0][-1])
            distance = tmp_col[0][-1] - tmp_col[0][0]
            midPoint = (i, int((tmp_col[0][0] + tmp_col[0][-1])/2))
            points.append((distance, (firstPoint, lastPoint), midPoint))
    # print("time process findpoints= {} ms".format((time.perf_counter_ns() - start1)/1000000))
    return points


def drawMid(img, array, numline = 20):
	# # draw line along image
    reser = []
    t = len(array)
    value = t // numline
    # value = int(value - value / 10)

    for i in range(t) :
        t = t - value

        cv.line(img, array[t][1][0], array[t][1][1], (255,t,0), 2)
        cv.circle(img, array[t][2], 5, (255,255,t), -1)

        if t < value :
            # reser.insert(0, (img.shape[1], array[-1][2][1]))
            # reser.append((0, array[0][2][1]))    
            reser.insert(0, (img.shape[1], array[-1][2][1]))
            reser.append((0, array[t+value][2][1]))                  
            break

        reser.append(array[t][2])

    return reser


def isDetermineBackAndBelly(subImageToDraw, subImageThreshold, onSiteValue = 0, onsite = False):
    verify = False
     #              determine back and belly                         #
    subImageCopy = subImageToDraw.copy()
    splitImage = subImageToDraw.copy()
    splitBinaryBellyImage = subImageThreshold.copy()

    if onsite == True:
        points = onSiteValue
    else:
        points = findPoints(subImageThreshold)

    aNumberOfLine = drawMid(subImageToDraw, points, numline= 9)
    

    #	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#
                                    # SPLIT IMAGE INTO 2 PARTS

    cv.polylines(subImageCopy, [np.array(aNumberOfLine)], isClosed= False, color= ( 0, 0, 0), thickness= 2) 

    backImage = splitImage.copy()
    bellyImage = splitImage.copy()
    newMidBack = aNumberOfLine.copy()
    newMidBelly = aNumberOfLine.copy()

    # # fillpoly apart top side
    newMidBelly.append((0,0))
    newMidBelly.append((subImageCopy.shape[1], 0))

    # # fillpoly apart bottom side
    newMidBack.append((0, subImageCopy.shape[0]))
    newMidBack.append((subImageCopy.shape[1], subImageCopy.shape[0]))
    
        # # back side and belly
    cv.fillPoly(backImage, pts= [np.array(newMidBack)], color= (0,0,0))  	# BACK
    cv.fillPoly(bellyImage, pts= [np.array(newMidBelly)], color= (0,0,0))	# BELLY

    # cv.imwrite('{}\\backImage{}.jpg'.format(pathSaveImage, id), backImage)
    # cv.imwrite('{}\\bellyImage{}.jpg'.format(pathSaveImage, id), bellyImage)

    #	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#

    #	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#

    h, w = backImage.shape[:2]

    backImage = backImage[0:h, int(0.5*w) : int(0.8*w)]
    bellyImage = bellyImage[0:h, int(0.5*w) : int(0.8*w)]

    backGray = cv.cvtColor(backImage, cv.COLOR_BGR2GRAY)
    bellyGray = cv.cvtColor(bellyImage, cv.COLOR_BGR2GRAY)

    backBinaryImageSub = cv.threshold(backGray, 0, 255, cv.THRESH_BINARY)[1]
    bellyBinaryImageSub = cv.threshold(bellyGray, 0, 255, cv.THRESH_BINARY)[1]

    meanBackGray = round(cv.mean(backGray, mask= backBinaryImageSub)[0],1)
    meanBellyGray = round(cv.mean(bellyGray, mask= bellyBinaryImageSub)[0],1)

    print("meanBackGray=", meanBackGray)
    print("meanBellyGray=", meanBellyGray)

    # cv.imwrite('D:\\Project\\d\\backImage.jpg', backImage)
    # cv.imwrite('D:\\Project\\d\\bellyImage.jpg', bellyImage)

    if meanBellyGray > meanBackGray :
        verify = True
        cv.fillPoly(splitBinaryBellyImage, pts= [np.array(newMidBelly)], color= (0,0,0))
        print('Continues proccessing then the next step is segment fin')
        
    return backImage, bellyImage, subImageCopy, verify, splitBinaryBellyImage


def parameterCalib():
    # cameraMatrix = np.array([[1.50392060e+03, 0.00000000e+00, 1.00861013e+03],
    #                          [0.00000000e+00, 1.50877936e+03, 5.83831826e+02],
    #                          [0.00000000e+00, 0.00000000e+00, 1.00000000e+00]])

    # Distortion_Parameters = np.array([[-0.15297373,  0.09234028,  0.00272974, -0.00363138,  0.00640345]])

    # cameraMatrix = np.array([[4.38387445e+03, 0.00000000e+00, 9.05271785e+02],
    #                          [0.00000000e+00, 4.53534042e+03, 5.74974349e+02],
    #                          [0.00000000e+00, 0.00000000e+00, 1.00000000e+00]])

    # Distortion_Parameters = np.array([[-1.40914232e+00,  6.79113169e+00,  1.46092441e-02, -3.77859355e-02, 6.40887570e+01]])


    cameraMatrix = np.array([[4.50120624e+03, 0.00000000e+00, 2.93651417e+02],
                            [0.00000000e+00, 4.76083681e+03, 6.67946500e+01],
                            [0.00000000e+00, 0.00000000e+00, 1.00000000e+00]]) 
    Distortion_Parameters = np.array([[-1.2567939,  -2.00653878,  0.16111217,  0.17518804, -0.19850368]])

    return cameraMatrix, Distortion_Parameters


def createMask():
    mask = cv.createBackgroundSubtractorMOG2(detectShadows= True)
    return mask


def splitFin(splitBelly2Draw, splitBinaryBelly):
    pointBelly = findPoints(splitBinaryBelly)
    aNumberOfLines = drawMid(splitBinaryBelly.copy(), pointBelly, numline= 40)
    

    #	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#
                                    # SPLIT IMAGE INTO 2 PARTS

    cv.polylines(splitBelly2Draw, [np.array(aNumberOfLines)], isClosed= False, color= ( 0, 0, 0), thickness= 1) 

    bellyImage = splitBelly2Draw.copy()
    newMidBelly = aNumberOfLines.copy()

    # # fillpoly apart top side
    newMidBelly.append((0,0))
    newMidBelly.append((bellyImage.shape[1], 0))
    
        # # side belly
    cv.fillPoly(bellyImage, pts= [np.array(newMidBelly)], color= (0,0,0))	# BELLY

    return bellyImage


def main():
    global n
    backSub = createMask()
    path = "D:\\video\\vd3.avi"
    video = cv.VideoCapture(path)

    while True:
        ret, img = video.read()

        try:
            img = cv.resize(img, None, fx=scale, fy=scale)

            madeImg= img.copy()
            originalImgshow = img.copy()
            maskImage = backSub.apply(img)


            #           determine width take a      #
            determineLocateLine(madeImg)


            #           take a picture          #
            saveImage(madeImg, originalImgshow, maskImage, pathSaveImage, True)


            # maskImage = np.stack((maskImage, )*3, axis= -1)
            # mergeImg = np.hstack((madeImg, maskImage))
            mergeImg = cv.resize(madeImg, (0,0), fx=0.5, fy= 0.5)

            cv.imshow('namespace', mergeImg)
            # cv.imshow('title', ORIIMG)

        except:
            pass

        k = cv.waitKey(1)
        if k == 27:
            break

        elif k == ord('s'):
            cv.imwrite('D:\Project\Take a picture\image{}.jpg'.format(n), madeImg)
            # cv.imwrite('D:\Project\imageCalib\image{}.jpg'.format(n), originalImgshow)  
            n += 1
            print("Image saved by press keyboard !")
        
    # Releasing the resource    
    video.release()

    cv.destroyAllWindows()


if __name__ == '__main__':
    cameraMatrix, Distortion_Parameters = parameterCalib()
    main()

