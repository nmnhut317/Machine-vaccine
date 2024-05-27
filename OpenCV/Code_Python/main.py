'''
A simple Program for grabing video from basler camera and converting it to opencv img.
Tested on Basler acA1300-200uc (USB3, linux 64bit , python 3.5)

'''
from pypylon import pylon
import cv2 as cv
import numpy as np
import glob
import pandas as pd
import time
import os
from datetime import datetime
import serial
from openpyxl import Workbook


wb = Workbook()
ws = wb.active
ser = serial.Serial('COM3', 115200)
ws.append(['ID','LENGTH (mm)', 'INJECTION (mm)', 'TIME'])

n = 0
id = 0
# thu nho anh voi ty le = 50% anh goc
rate_scale = 0.5

collums = 2

lengthFish = 0
positionCurrent = 0
positionAhead = 0

currentTime = datetime.now()

# duong dan luu tru file
pathSaveImage = 'e'
pathSaveData = 'D:/Project/data.xlsx'

# kiem tra duong dan file
CHECK_DIR = os.path.isdir(pathSaveImage)
# if directory does not exist create
if not CHECK_DIR:
    os.makedirs(pathSaveImage)
    print(f'"{pathSaveImage}" Directory is created')
else:
    print(f'"{pathSaveImage}" Directory already Exists.') 

# scale hinh anh
def scale(img, cnt):
        h, w = img.shape[:2]
        scaledH = int(h*cnt)
        scaledW = int(w*cnt)
        dim = (scaledW, scaledH)
        res = cv.resize(img, dim, interpolation = cv.INTER_AREA)
        return res


def image_resize(image, width = None, height = None, inter = cv.INTER_AREA):
    # initialize the dimensions of the image to be resized and
    # grab the image size
    dim = None
    (h, w) = image.shape[:2]

    # if both the width and height are None, then return the
    # original image
    if width is None and height is None:
        return image

    # check to see if the width is None
    if width is None:
        # calculate the ratio of the height and construct the
        # dimensions
        r = height / float(h)
        dim = (int(w * r), height)

    # otherwise, the height is None
    else:
        # calculate the ratio of the width and construct the
        # dimensions
        r = width / float(w)
        dim = (width, int(h * r))

    # resize the image
    resized = cv.resize(image, dim, interpolation = inter)

    # return the resized image
    return resized


# tinh chieu dai anh ca
def drawCenterLine(img2Draw, binaryImage):
    point = findPoints(binaryImage)
    # thay doi so luong diem chia tren anh
    site = [0, 0.05, 0.10, 0.20, 0.35, 0.5, 0.90, 1]
    pointSite = []
    siteLength = len(point)

    for i in range(siteLength):
        t = int(site[i]*img2Draw.shape[1] - 2)

        if t < 0:
            t = 0

        pointSite.append(point[t][2])

        if site[i] == 1:
            cv.circle(img2Draw, point[t][2], 2, (255,255,t), -1) #11/22
            break

        cv.line(img2Draw, point[t][1][0], point[t][1][1], (255,255,255), 1)
        cv.circle(img2Draw, point[t][2], 2, (255,255,t), -1)

    cv.polylines(img2Draw, [np.array(pointSite)], isClosed= False, color= (255, 255, 255), thickness= 1) 

    return pointSite, point


# tinh chieu dai anh ca
def distanceMultiplePoints(arrayPoint):

    datatype = object
    arrayPoint = np.array(arrayPoint, dtype= datatype)
    # print("after length",arrayPoint, len(arrayPoint))
    # priviousPoint = (arrayPoint[0][0], arrayPoint[0][1])
    priviousPoint = arrayPoint[0]
    dist = 0
    numOfArray = len(arrayPoint)
    # print("length", numOfArray)

    for i in range(numOfArray):
        currentPoint = arrayPoint[i]
        dist = dist + np.linalg.norm(priviousPoint - currentPoint)
        priviousPoint = currentPoint

    dist = round(dist, 1)

    return dist


# ham so xac dinh diem tiem
def convertLenghToInjection(length):
    positionInjection = int((length*0.3362 + 2.6474)*0.63)
    return positionInjection


def isCheckDirection():
    global positionCurrent, positionAhead
    if positionCurrent < positionAhead :
        positionDynamic = positionAhead - positionCurrent
        transmitData(0, positionDynamic) # 0 tien
        print("position= ", positionDynamic)

    if positionCurrent > positionAhead:
        positionDynamic = abs(positionAhead - positionCurrent)
        transmitData(1, positionDynamic) # 1 lui
        print("position= ", positionDynamic)

    if positionCurrent == positionAhead:
        transmitData(0, 0)


# truyen du lieu uart den vi dieu khien
def transmitData(direction, position):
    value = bytes('{} {}\n'.format(direction, position), encoding= 'utf8')
    ser.write(value)
    print(value)


def transmitPosition(position, check):
    if(check == True):
        value = bytes('{}\n'.format(position), encoding= 'utf8')
        
    else:
        value = bytes('\r', encoding= 'utf8')

    ser.write(value)    
    print(value)
          

# chong bien dang anh
def calibInputImage(img, cameraMatrix, Distortion_Parameters, path, id):
    h, w = img.shape[:2]
    newcameramtx, roi = cv.getOptimalNewCameraMatrix(cameraMatrix, Distortion_Parameters, (w,h), 1, (w,h))
    # undistort
    Distortion = cv.undistort(img, cameraMatrix, Distortion_Parameters, None, newcameramtx)
    # crop the image
    x, y, w, h = roi
    Distortion = Distortion[y:y+h, x:x+w]
    # cv.imwrite('{}/calibresult{}.jpg'.format(path, id), Distortion)
    # outputImage= cv.imread('{}/calibresult{}.jpg'.format(path, id))

    return Distortion


def findContour(threshold):
    gray = cv.cvtColor(threshold, cv.COLOR_BGR2GRAY)
    threshold = cv.threshold(gray, 0, 255, cv.THRESH_BINARY)[1]
    contour = cv.findContours(threshold, cv.RETR_EXTERNAL, cv.CHAIN_APPROX_SIMPLE)[0]

    if len(contour)>0:
        contour = max(contour, key= cv.contourArea) 
        rect = cv.minAreaRect(contour)

    return rect 


# tinh chieu dai va vi tri tiem tren ca
def processImageByCenterLine(readImage, path, id, printValue= False):
    global positionCurrent, positionAhead, collums, rate_scale  
    img= calibInputImage(readImage, cameraMatrix, Distortion_Parameters, path, id)
    newImg= np.zeros(img.shape, dtype= np.uint8)

    hsv = cv.cvtColor(img, cv.COLOR_BGR2HSV)
    # thay doi gia tri HSV phu hop
    low_hsv = np.array([68, 116, 95])
    high_hsv = np.array([179, 255, 255])

    mask = cv.inRange(hsv, low_hsv, high_hsv)
    mask = ~mask
    kernel = np.ones((5,5),np.uint8)
    mask = cv.erode(mask, kernel)
    mask = cv.medianBlur(mask, 5)
    # mask = cv.morphologyEx(mask, cv.MORPH_CLOSE, kernel, iterations= 1)
    contour = cv.findContours(mask, cv.RETR_EXTERNAL, cv.CHAIN_APPROX_SIMPLE)[0]

    if len(contour) > 0:
        contour = max(contour, key= cv.contourArea) 

        #	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#
        #                   DRAW A NEW IMAGE            #
        cv.drawContours(newImg, [contour], 0, (255, 255, 255), -1)
        rect = findContour(newImg)

        #	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#
        #222, 1283
        subImg = getSubImage(rect, newImg, img)[0]

        newMask = cv.split(newImg)[0]
        result = cv.bitwise_and(img, img, mask= newMask)
        subColor =  getSubImage(rect, result, img)[0]

        subCenterLine = subColor.copy()
        pointSite, onsitePoints = drawCenterLine(subCenterLine, subImg)
        dist = distanceMultiplePoints(pointSite)
        # print("length pixel= {dist} (pixel)")
        lengthFish = calculatorLengthByCenterLine((1/rate_scale)*(dist+6), writeLength= True)
        lengthFish = int(lengthFish)

        cv.imwrite('{}\\newImage{}.jpg'.format(path, id), result)
        cv.imwrite('{}\\newConvertImage{}.jpg'.format(path, id), subColor)

        drawInjectionImage = subColor.copy()

        back, belly, midLine, check, splitBellyBinary = isDetermineBackAndBelly(subColor, subImg, onsitePoints, onsite= True)

        borderImage = None #11/22

        if check == True:
            #	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	# 
            positionCurrent = convertLenghToInjection(lengthFish)
            injection_pixel = int((positionCurrent / 0.17)/(1/rate_scale)) 
            transmitPosition(positionCurrent, True)
            print(f'injection at = {positionCurrent}(mm)')
            # print(f'\ninjection at = {positionCurrent}(mm) the same with {injection_pixel*(1/rate_scale)}(pixel)')
            injection_OnImgX = int(drawInjectionImage.shape[1] - injection_pixel)
            cv.circle(drawInjectionImage, (injection_OnImgX, onsitePoints[injection_OnImgX][1][1][1] - 2), 10, (0,0,255), 2)
            # 11/27
            borderImage = cv.copyMakeBorder(drawInjectionImage, 50, 100, 50, 50, cv.BORDER_CONSTANT, value=[0,0,0])
            cv.putText(borderImage, f'Injection position={positionCurrent}(mm)', (20, borderImage.shape[0] - 50), cv.FONT_HERSHEY_COMPLEX, 1, (255,255,255))
            # 11/27
            cv.imwrite('{}\\borderImage{}.jpg'.format(path, id), borderImage)

        # when wrong side
        if check == False:
            transmitPosition(positionCurrent, False)
            # 11/27
            borderImage = cv.copyMakeBorder(drawInjectionImage, 50, 50, 50, 50, cv.BORDER_CONSTANT, value=[0,0,0])
            #	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#

        #11/22
        cv.putText(borderImage, f'Length fish={lengthFish}(mm)', (20, borderImage.shape[0] - 20), cv.FONT_HERSHEY_COMPLEX, 1, (255,255,255))

        # write data into excel
        ws[f'A{collums}'].value = id
        ws[f'B{collums}'].value = lengthFish
        ws[f'C{collums}'].value = positionCurrent
        ws[f'D{collums}'].value = currentTime
        collums += 1
        wb.save(pathSaveData)

        print("time process= {}(ms)".format(int((time.perf_counter_ns() - start)/1000000)))
        
        # colorBelly = splitFin(belly.copy(), splitBellyBinary)

        result = cv.resize(result, (0,0), fx=0.5, fy=0.5)
        subColor = cv.resize(subColor, (0,0), fx= 0.5, fy= 0.5)

        mergeB2B = np.hstack((back, belly))
        mergeB2B = cv.resize(mergeB2B, (0,0), fx= 0.5, fy= 0.5)

        displayInjection = cv.resize(drawInjectionImage, (0,0), fx= 0.5, fy= 0.5)
        displayBorderImage = cv.resize(borderImage, (0,0), fx= 1, fy= 1)

        mergeCenter = np.hstack((midLine, subCenterLine))
        mergeCenter = cv.resize(mergeCenter, (0,0), fx= 0.5, fy= 0.5)

        # cv.imshow("new image", result)
        # cv.imshow("displayInjection", displayInjection)
        cv.imshow("displayBorderImage", displayBorderImage)

        # cv.imshow("back and belly", mergeB2B)
        # cv.imshow("subColor", mergeCenter)
        # cv.imshow("subimage", subCenterLine)
   

def findThreshold(img):
    contour = cv.findContours(img, cv.RETR_EXTERNAL, cv.CHAIN_APPROX_SIMPLE)[0]

    if len(contour) > 0:
        contour = sorted(contour, key= cv.contourArea, reverse= True)
        contour = contour[0]
        rects = cv.minAreaRect(contour) 

    return rects


# luu anh khi co doi tuong
def saveImage(drawframe, img, threshold, path, takeImage = False):
    global id, start

    contour = cv.findContours(threshold, cv.RETR_EXTERNAL, cv.CHAIN_APPROX_SIMPLE)[0]
    
    try:
        contour = max(contour, key= cv.contourArea)

        if len(contour) > 0:
                x, y, w, h = cv.boundingRect(contour)
                CX = int((x+w)/2)
                CY = int((y+h)/2)
   
                if cv.contourArea(contour) > 10000 and takeImage == True:
                    cv.circle(drawframe, (CX , CY), 2, (0,0,255), 15)
                    cv.putText(drawframe, 'CX= {} CY= {}'.format(CX , CY), (CX , CY), cv.FONT_HERSHEY_COMPLEX, 1, 255, 2)                    
                    # vung luu anh tu (giua anh - x va giua anh - y)
                    if CX > int(img.shape[1]/2) - 90  and CX < int(img.shape[1]/2) - 30:
                        start = time.perf_counter_ns()
                        id += 1
                        print("\nimage saved")
                        print("\ninto readImage")
                        try:
                            processImageByCenterLine(img, path, id, True)
                            print("\nout readImage")
                        except:
                            pass
   
    except:
        pass


# chuyen chieu dai pixel to mm
def calculatorLengthByCenterLine(lengthPixel, lengthFrame = 340, widthPixelFrame = 1920, writeLength = False):

    length = lengthPixel* 0.17

    if writeLength == True:
        print(f"length = {int(length)}(mm)")

    return length


def calibPylon(cameraMatrix, dist, img):
    # Undistortion  
    h, w = img.shape[:2]
    newcameramtx, roi = cv.getOptimalNewCameraMatrix(cameraMatrix, dist, (w,h), 1, (w,h))
    # undistort
    dst = cv.undistort(img, cameraMatrix, dist, None, newcameramtx)

    return dst


# xac dinh vung luu anh
def determineLocateLine(img):
    cv.line(img, (int(img.shape[1]/2) - 30, 0), (int(img.shape[1]/2) - 30, img.shape[0]), (0,0,255), 2)
    cv.line(img, (int(img.shape[1]/2) - 90, 0), (int(img.shape[1]/2) - 90, img.shape[0]), (0,255,0), 2)


# lay khung anh bao quanh doi tuong
def getSubImage(rect, src, img, drawPolyline = False):
        
    box = cv.boxPoints(rect)
    box = np.intp(box)
        
    width = int(rect[1][0])
    height = int(rect[1][1])

    
    theta = rect[2]

    if drawPolyline == True:
        cv.polylines(img, [box], True, (0,0,255), 3)

    src_pts = box.astype("float32")
    dst_pts = np.array([[0, height-1],
                                [0, 0],
                                [width-1, 0],
                                [width-1, height-1]], dtype="float32")

    M = cv.getPerspectiveTransform(src_pts, dst_pts)
    warped = cv.warpPerspective(src, M, (width, height))

    if width < height :
        warped = cv.rotate(warped, cv.ROTATE_90_CLOCKWISE)   
                
    return warped, width  


def calibCamera(chessboardSize, frameSize, pathImage, saveImage = False):

    # termination criteria
    criteria = (cv.TERM_CRITERIA_EPS + cv.TERM_CRITERIA_MAX_ITER, 30, 0.001)

    # prepare object points, like (0,0,0), (1,0,0), (2,0,0) ....,(6,5,0)
    objp = np.zeros((chessboardSize[0]*chessboardSize[1], 3), np.float32)
    objp[:,:2] = np.mgrid[0:chessboardSize[0], 0: chessboardSize[1]].T.reshape(-1,2)

    # Arrays to store object points and image points from all the images.
    objpoints = [] # 3d point in real world space
    imgpoints = [] # 2d points in image plane.

    images = glob.glob('{}/*.jpg'.format(str(pathImage)))

    a_number_of_images = 1

    for fname in images:

        img = cv.imread(fname)

        gray = cv.cvtColor(img, cv.COLOR_BGR2GRAY)

        # Find the chess board corners
        ret, corners = cv.findChessboardCorners(gray, chessboardSize, None)

        # If found, add object points, image points (after refining them)
        if ret == True:
            objpoints.append(objp)
            corners2 = cv.cornerSubPix(gray,corners, (11,11), (-1,-1), criteria)
            imgpoints.append(corners2)

            # Draw and display the corners
            cv.drawChessboardCorners(img, chessboardSize, corners2, ret)

            h, w = img.shape[:2]
            print(img.shape)
            img = cv.resize(img, (int(w/2), int(h/2)), fx= 0.5, fy= 0.5)
            
            # cv.imshow('img', img)

            # cv.waitKey(1000)

        a_number_of_images += 1

    print('a_number_of_images:', a_number_of_images)

    cv.destroyAllWindows()

    # calibration
    ret, cameraMatrix, dist, rvecs, tvecs = cv.calibrateCamera(objpoints, imgpoints, frameSize, None, None)
    # print("Camera Calibrate:", ret)
    print("\ncameraMatrix", cameraMatrix)
    print("\nDistortion Parameters:", dist)
    # print("\nRotation Vector:", rvecs)
    # print("\nTranstation Vector:", tvecs)

    # pickle.dump((cameraMatrix, dist), open('calibration.pkl', "wb"))
    # pickle.dump(cameraMatrix, open('cameraMatrix.pkl', "wb"))
    # pickle.dump(dist, open('calibration.pkl', "wb"))

    k = 1
    #               the same feature            #
    # THE FIRST METHOD
    if saveImage == True:
        # Undistortion  
        img = cv.imread('{}\Image1.png'.format(pathImage))
        h, w = img.shape[:2]
        newcameramtx, roi = cv.getOptimalNewCameraMatrix(cameraMatrix, dist, (w,h), 1, (w,h))
        # undistort
        dst = cv.undistort(img, cameraMatrix, dist, None, newcameramtx)
        # crop the image
        x, y, w, h = roi
        dst = dst[y:y+h, x:x+w]
        cv.imwrite('{}\calibresult{}.png'.format(pathImage, k), dst)
        k += 1

    # check error
    mean_error = 0

    for i in range(len(objpoints)):
            imgpoints2, _ = cv.projectPoints(objpoints[i], rvecs[i], tvecs[i], cameraMatrix, dist)
            error = cv.norm(imgpoints[i], imgpoints2, cv.NORM_L2)/len(imgpoints2)
            mean_error += error

    print( "total error: {}".format(mean_error/len(objpoints)) ) 
        
    return cameraMatrix, dist


def findPoints(img, indxStart = 0, indxEnd = 1):
    # start1 = time.perf_counter_ns()
    # find smalles value in each column
    origi = int(img.shape[1])
    midPt = int(img.shape[1] / 2)
    firstPt = int(img.shape[1] /7)
    limitPt = int(img.shape[1] - img.shape[1] /10)
    coordinate = []
    disPoint = []
    points = []

    indxStart= int(origi*indxStart)
    indxEnd = int(origi*indxEnd)


    # for i in range(img.shape[1]):
    for i in range(indxStart, indxEnd):
    # for i in range(midPt, limitPt):
        tmp_col = np.where((img[:, i] == 255))

        if len(tmp_col[0]) > 2:
            firstPoint =  (i, tmp_col[0][0])
            lastPoint =  (i, tmp_col[0][-1])
            distance = tmp_col[0][-1] - tmp_col[0][0]
            midPoint = (i, int((tmp_col[0][0] + tmp_col[0][-1])/2))
            points.append((distance, (firstPoint, lastPoint), midPoint))
    # print("time process findpoints= {} ms".format((time.perf_counter_ns() - start1)/1000000))
    return points

# xac dinh diem chia doc than ca
def drawMid(img, array, numline = 20):
	# # draw line along image
    reser = []
    t = len(array)
    value = t // numline

    for i in range(t) :
        t = t - value

        cv.line(img, array[t][1][0], array[t][1][1], (255,t,0), 2)
        cv.circle(img, array[t][2], 5, (255,255,t), -1)

        if t < value :
            # reser.insert(0, (img.shape[1], array[-1][2][1]))
            # reser.append((0, array[0][2][1]))    
            reser.insert(0, (img.shape[1], array[-1][2][1]))
            reser.append((0, array[t+value][2][1]))                  
            break

        reser.append(array[t][2])

    return reser


# xac dinh mac bung ca
def isDetermineBackAndBelly(subImageToDraw, subImageThreshold, onSiteValue = 0, onsite = False):
    verify = False
     #              determine back and belly                         #
    subImageCopy = subImageToDraw.copy()
    splitImage = subImageToDraw.copy()
    splitBinaryBellyImage = subImageThreshold.copy()

    if onsite == True:
        points = onSiteValue
    else:
        points = findPoints(subImageThreshold)

    aNumberOfLine = drawMid(subImageToDraw, points, numline= 9)
    

    #	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#
                                    # SPLIT IMAGE INTO 2 PARTS

    cv.polylines(subImageCopy, [np.array(aNumberOfLine)], isClosed= False, color= ( 0, 0, 0), thickness= 1) 

    backImage = splitImage.copy()
    bellyImage = splitImage.copy()
    newMidBack = aNumberOfLine.copy()
    newMidBelly = aNumberOfLine.copy()

    # # fillpoly apart top side
    newMidBelly.append((0,0))
    newMidBelly.append((subImageCopy.shape[1], 0))

    # # fillpoly apart bottom side
    newMidBack.append((0, subImageCopy.shape[0]))
    newMidBack.append((subImageCopy.shape[1], subImageCopy.shape[0]))
    
        # # back side and belly
    cv.fillPoly(backImage, pts= [np.array(newMidBack)], color= (0,0,0))  	# BACK
    cv.fillPoly(bellyImage, pts= [np.array(newMidBelly)], color= (0,0,0))	# BELLY


    #	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#

    
    # hBack, wBack = backImage.shape[:2]
    # backImageSub = backImage[0:hBack, int(0.5*wBack):wBack]
    # rectBack = findContour(backImageSub)
    # backImageSub = getSubImage(rectBack, backImageSub, backImage.copy())[0]
    # backGray = cv.cvtColor(backImageSub, cv.COLOR_BGR2GRAY)
    # backBinaryImageSub = cv.threshold(backGray, 0, 255, cv.THRESH_BINARY)[1]
    # # cv.imshow("backImage", backImageSub)
    # # cv.imshow("backImages", backBinaryImageSub)
    # # print("size backImage", backImageSub.shape)

    
    # hBelly, wBelly = bellyImage.shape[:2]
    # bellyImageSub = bellyImage[0:hBelly, int(0.5*wBelly):wBelly]
    # rectBelly = findContour(bellyImageSub)
    # bellyImageSub = getSubImage(rectBelly, bellyImageSub, bellyImage.copy())[0]
    # bellyGray = cv.cvtColor(bellyImageSub, cv.COLOR_BGR2GRAY)
    # bellyBinaryImageSub = cv.threshold(bellyGray, 0, 255, cv.THRESH_BINARY)[1]
    # # cv.imshow("bellyImage", bellyImageSub)
    # # cv.imshow("bellyImages", bellyBinaryImageSub)
    # # print("size bellyImage", bellyImageSub.shape)
    

    # meanBackGray = round(cv.mean(backGray, mask= backBinaryImageSub)[0],1)
    # meanBellyGray = round(cv.mean(bellyGray, mask= bellyBinaryImageSub)[0],1)

    #	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#

    h, w = backImage.shape[:2]

    backImage = backImage[0:h, int(0.5*w):w]
    bellyImage = bellyImage[0:h, int(0.5*w):w]

    backGray = cv.cvtColor(backImage, cv.COLOR_BGR2GRAY)
    bellyGray = cv.cvtColor(bellyImage, cv.COLOR_BGR2GRAY)

    backBinaryImageSub = cv.threshold(backGray, 0, 255, cv.THRESH_BINARY)[1]
    bellyBinaryImageSub = cv.threshold(bellyGray, 0, 255, cv.THRESH_BINARY)[1]

    meanBackGray = round(cv.mean(backGray, mask= backBinaryImageSub)[0],1)
    meanBellyGray = round(cv.mean(bellyGray, mask= bellyBinaryImageSub)[0],1)

    print("meanBackGray=", meanBackGray)
    print("meanBellyGray=", meanBellyGray)

    # # cv.imwrite('D:\\Project\\b\\image.jpg', backImage)

    if meanBellyGray > meanBackGray :
        verify = True
        cv.fillPoly(splitBinaryBellyImage, pts= [np.array(newMidBelly)], color= (0,0,0))
        print('continues proccessing then the next step is segment fin')
        
    return backImage, bellyImage, subImageCopy, verify, splitBinaryBellyImage


# xac dinh vay cua ca
def splitFin(splitBelly2Draw, splitBinaryBelly):
    pointBelly = findPoints(splitBinaryBelly)
    aNumberOfLines = drawMid(splitBinaryBelly.copy(), pointBelly, numline= 40)
    

    #	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#	#
                                    # SPLIT IMAGE INTO 2 PARTS

    cv.polylines(splitBelly2Draw, [np.array(aNumberOfLines)], isClosed= False, color= ( 0, 0, 0), thickness= 1) 

    bellyImage = splitBelly2Draw.copy()
    newMidBelly = aNumberOfLines.copy()

    # # fillpoly apart top side
    newMidBelly.append((0,0))
    newMidBelly.append((bellyImage.shape[1], 0))
    
        # # side belly
    cv.fillPoly(bellyImage, pts= [np.array(newMidBelly)], color= (0,0,0))	# BELLY

    return bellyImage


# thong so calib camera
def parameterCalib():
    # (1600, 800)
    # total_error = 0.0355667129386486
    # cameraMatrix = np.array([[1.50392060e+03, 0.00000000e+00, 1.00861013e+03],
    #                          [0.00000000e+00, 1.50877936e+03, 5.83831826e+02],
    #                          [0.00000000e+00, 0.00000000e+00, 1.00000000e+00]])

    # Distortion_Parameters = np.array([[-0.15297373,  0.09234028,  0.00272974, -0.00363138,  0.00640345]])
    
    # (800, 400)
    # total_error = 0.0619987917507736
    cameraMatrix = np.array([[4.50120624e+03, 0.00000000e+00, 2.93651417e+02],
                            [0.00000000e+00, 4.76083681e+03, 6.67946500e+01],
                            [0.00000000e+00, 0.00000000e+00, 1.00000000e+00]]) 
    Distortion_Parameters = np.array([[-1.2567939,  -2.00653878,  0.16111217,  0.17518804, -0.19850368]])

    # (400, 200)
    # total_error =  0.06370957814914902
    # cameraMatrix = np.array([[5.85290209e+03, 0.00000000e+00, 1.92736517e+02],
    #                         [0.00000000e+00, 6.47323948e+03, 7.70371238e+01],
    #                         [0.00000000e+00, 0.00000000e+00, 1.00000000e+00]])

    # Distortion_Parameters = np.array([[ -2.02498813, -10.23408505,   0.20129337,   0.25039514,  59.03975702]])
    
    return cameraMatrix, Distortion_Parameters


# mask cho doi tuong di chuyen
def createMask():
    mask = cv.createBackgroundSubtractorMOG2(detectShadows= True)
    # queue.put(mask)
    return mask


if __name__ =="__main__":

    # cameraMatrix, Distortion_Parameters = calibCamera((15,9), (1920, 1200), 'a', False)

    cameraMatrix, Distortion_Parameters = parameterCalib()


    backSub = createMask()


    # conecting to the first available camera
    camera = pylon.InstantCamera(pylon.TlFactory.GetInstance().CreateFirstDevice())

    # Grabing Continusely (video) with minimal delay
    camera.StartGrabbing(pylon.GrabStrategy_LatestImageOnly) 
    converter = pylon.ImageFormatConverter()

    # converting to opencv bgr format
    converter.OutputPixelFormat = pylon.PixelType_BGR8packed
    converter.OutputBitAlignment = pylon.OutputBitAlignment_MsbAligned

    # Set the upper limit of the camera's frame rate to 30 fps
    camera.AcquisitionFrameRateEnable.SetValue(True)
    camera.AcquisitionFrameRate.SetValue(60.0)
    print("FPS:", camera.ResultingFrameRate.GetValue())

    # camera.ExposureTimeMode.SetValue(ExposureTimeMode_Standard)
    # camera.ExposureTime.SetValue(camera.ExposureTime.Min)
    camera.ExposureTime.SetValue(1000)

    # camera.Width.SetValue(1600)
    # # Set the height to 500
    # camera.Height.SetValue(800)
    # # Set the offset to 0
    # camera.OffsetX.SetValue(200)
    # camera.OffsetY.SetValue(100)

    while camera.IsGrabbing():
        grabResult = camera.RetrieveResult(5000, pylon.TimeoutHandling_ThrowException)
    
        if grabResult.GrabSucceeded():
            # Access the image data
            image = converter.Convert(grabResult)
            img = image.GetArray()
            # cv.namedWindow('title', cv.WINDOW_NORMAL)
            # ORIIMG = img.copy()

            # pylonFrame = calibPylon(cameraMatrix, Distortion_Parameters, img)    
            img= scale(img, rate_scale)       # original (1600, 800)
            madeImg= img.copy()
            originalImgshow = img.copy()
            maskImage = backSub.apply(img)


            #           determine width take a      #
            determineLocateLine(madeImg)


            #           take a picture          #
            saveImage(madeImg, originalImgshow, maskImage, pathSaveImage, True)


            # maskImage = np.stack((maskImage, )*3, axis= -1)
            # mergeImg = np.hstack((madeImg, maskImage))
            # mergeImg = cv.resize(mergeImg, (0,0), fx=0.5, fy= 0.5)

            mergeImg = cv.resize(madeImg, (0,0), fx=1, fy= 1)   #11/27

            cv.imshow('Frame work', mergeImg)   #11/27
            # cv.imshow('title', ORIIMG)

            # cv.setMouseCallback('namespace', click_event)
            k = cv.waitKey(1)
            if k == 27:
                break

            elif k == ord('s'):
                cv.imwrite('D:\Project\Take a picture\image{}.jpg'.format(n), originalImgshow)
                # cv.imwrite('D:\Project\imageCalib\image{}.jpg'.format(n), originalImgshow)  
                n += 1
                print("Image saved by press keyboard !")

        grabResult.Release()
        
    # Releasing the resource    
    camera.StopGrabbing()

    cv.destroyAllWindows()

